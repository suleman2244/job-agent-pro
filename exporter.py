import pandas as pd
import os
from config import OUTPUT_FILENAME

def export_to_excel(jobs_list):
    """
    Exports a list of job dictionaries to a professional, formatted Excel file.
    Includes clickable HYPERLINKS and clean data categorization.
    """
    if not jobs_list:
        print("No jobs to export.")
        return None

    try:
        # Create DataFrame
        df = pd.DataFrame(jobs_list)
        
        # Ensure correct column order
        columns = ['title', 'company', 'location', 'link', 'emails', 'source']
        for col in columns:
            if col not in df.columns:
                df[col] = ""
        
        df = df[columns]

        # Clean strings for Excel
        def clean_val(val):
            if isinstance(val, list):
                return "\n".join(val) # Use newline for cleaner email listing
            if not isinstance(val, str):
                return val
            return "".join(c for c in val if c.isprintable() or c in "\n\r\t")

        for col in df.columns:
            df[col] = df[col].apply(clean_val)

        # Save using openpyxl with advanced formatting
        with pd.ExcelWriter(OUTPUT_FILENAME, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Job Leads')
            
            workbook = writer.book
            worksheet = writer.sheets['Job Leads']
            
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            
            # --- PROFESSIONAL STYLING ---
            header_font = Font(bold=True, color="FFFFFF", size=12)
            header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
            center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
            left_wrap = Alignment(horizontal='left', vertical='top', wrap_text=True)
            
            # Style header row
            for cell in worksheet[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_align
                
            # --- CLICKABLE LINKS & ROW STYLING ---
            link_font = Font(color="0563C1", underline="single")
            
            for row_idx, row in enumerate(worksheet.iter_rows(min_row=2, max_row=worksheet.max_row), start=2):
                # Make link clickable using Formula
                link_cell = worksheet.cell(row=row_idx, column=4) # Column 'link'
                raw_url = link_cell.value
                if raw_url and raw_url.startswith("http"):
                    # Use HYPERLINK formula for actual clickability
                    link_cell.value = f'=HYPERLINK("{raw_url}", "OPEN JOB POSTING")'
                    link_cell.font = link_font
                
                # Align other cells
                for cell in row:
                    if cell.column != 4: # Don't overwrite the link alignment if we want it specific
                        cell.alignment = left_wrap

            # --- AUTO-ADJUST COLUMN WIDTHS ---
            column_widths = {
                'A': 40, # Title
                'B': 25, # Company
                'C': 20, # Location
                'D': 25, # Link
                'E': 35, # Emails
                'F': 15  # Source
            }
            for col, width in column_widths.items():
                worksheet.column_dimensions[col].width = width

            # Freeze panes
            worksheet.freeze_panes = 'A2'
            
        print(f"Successfully exported {len(jobs_list)} jobs to {OUTPUT_FILENAME}")
        return os.path.abspath(OUTPUT_FILENAME)
    except Exception as e:
        print(f"Error exporting to Excel: {e}")
        return None
